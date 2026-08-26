"""Generate seed/cts_dummy_data.xlsx for CTS dummy QA data.

Review the Overview sheet before any agent loads this into Django or the app.
1500 commuters must be bulk-imported — do not create them one-by-one in Flutter.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("cts_dummy_data.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
HEADER_FONT = Font(color="FFC107", bold=True)
SECTION_FONT = Font(bold=True, size=14)
WARN_FILL = PatternFill("solid", fgColor="FFF3CD")
OK_FILL = PatternFill("solid", fgColor="D4EDDA")

PASSWORD = "password"
ADMIN_MOBILE = "7069036462"
DRIVER_COUNT = 10
COMMUTER_COUNT = 1500
UG_COUNT = 750
PG_COUNT = 750
STOP_NAMES = [
    "Campus Gate",
    "Railway Station",
    "City Bus Stand",
    "Market Circle",
    "Civil Hospital",
    "Housing Colony",
    "Highway Crossing",
    "College Main Gate",
]
POPS_PER_ROUTE = len(STOP_NAMES)  # was 3; richer dummy coverage per route
CAB_CAPACITY = 53  # live trip seats; roster is larger because not everyone comes
ROSTER_PER_BATCH = COMMUTER_COUNT // DRIVER_COUNT  # 150 assigned
COMING_PER_BATCH = 15  # isComing=true; must be <= CAB_CAPACITY

DRIVER_BASE = 9876544111
COMMUTER_BASE = 9876556701


def style_header(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, min_w=10, max_w=42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_w
        for cell in col[:80]:
            if cell.value is None:
                continue
            width = max(width, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def write_rows(ws, headers, rows) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    style_header(ws, len(headers))
    autosize(ws)


def overview_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "CTS dummy data — REVIEW BEFORE EXECUTE"
    ws["A1"].font = SECTION_FONT
    ws.merge_cells("A1:D1")

    lines = [
        [],
        ["Status", "PLAN ONLY — do not load until the user approves this workbook."],
        ["File", str(OUT)],
        ["Org admin login", ADMIN_MOBILE],
        ["Shared password", PASSWORD],
        [],
        ["Counts", "Value", "Notes"],
        ["Admins", 1, "Existing/org account. Do not POST a second ADMIN unless missing."],
        ["Drivers", DRIVER_COUNT, f"First phone {DRIVER_BASE}, then +1"],
        ["Cabs", DRIVER_COUNT, f"Capacity {CAB_CAPACITY} seats (not the full roster)"],
        ["Routes", DRIVER_COUNT, "R01 … R10"],
        ["POPs", DRIVER_COUNT * POPS_PER_ROUTE, f"{POPS_PER_ROUTE} stops per route, inLine 1–{POPS_PER_ROUTE}"],
        ["Batches", DRIVER_COUNT, "One morning+return slot per driver/cab/route"],
        ["Commuters", COMMUTER_COUNT, f"{UG_COUNT} UG + {PG_COUNT} PG"],
        ["Roster per batch", ROSTER_PER_BATCH, "Assigned riders; most stay home (isComing false)"],
        ["isComing=true", DRIVER_COUNT * COMING_PER_BATCH, f"{COMING_PER_BATCH} per batch for D2D — fits in {CAB_CAPACITY} seats"],
        [],
        ["Assignment rule", "Detail"],
        ["Commuter → batch", f"index % 10 → Batch-01 … Batch-10 ({ROSTER_PER_BATCH} roster each)"],
        ["Commuter → cab/driver/route", "Same index as batch (Driver 1 ↔ Cab GJ-06-D-1001 ↔ R01 ↔ Batch-01)"],
        ["Commuter → POP", f"Rotate {POPS_PER_ROUTE} POPs on that route: " + " / ".join(STOP_NAMES)],
        ["College / name", "UG1…UG750 then PG1…PG750 as username + collegeName UG or PG"],
        ["Commuter mobiles", f"{COMMUTER_BASE} … {COMMUTER_BASE + COMMUTER_COUNT - 1}"],
        ["Driver mobiles", f"{DRIVER_BASE} … {DRIVER_BASE + DRIVER_COUNT - 1}"],
        ["Email", "{username.lower()}@cts.test ; empty address → copy email (A3)"],
        ["QA driver to log in", f"{DRIVER_BASE} / {PASSWORD} (Driver 1, Batch-01)"],
        ["QA commuter to log in", f"{COMMUTER_BASE} / {PASSWORD} (UG1, Batch-01, isComing=true)"],
        [],
        ["Load order (backend bulk, NOT Flutter forms)", ""],
        ["1", "Login as admin in the app only to verify — do not hand-create 1500 rows"],
        ["2", "Create Routes R01–R10"],
        ["3", f"Create {POPS_PER_ROUTE} POPs per route"],
        ["4", f"Create Cabs (capacity {CAB_CAPACITY} — trip seats, not roster size)"],
        ["5", "Create Drivers and attach cab"],
        ["6", "Create Batches (attach driver)"],
        ["7", "Bulk-create 1500 commuters (Django / script / API). Flutter UI is too slow."],
        ["8", "Smoke QA: admin emulator + driver phone (see docs/TESTING.md)"],
        [],
        ["Do not", "Wipe existing live org data without explicit user OK"],
        ["Do not", f"Mark more than {CAB_CAPACITY} isComing per batch (cab seats)"],
        ["Do not", "Set all 1500 isComing=true (live WS queue would be huge)"],
        ["Do not", "Use 10.0.2.2 in .env while the phone is in the same test"],
    ]
    for row in lines:
        ws.append(row)
    ws["A3"].fill = WARN_FILL
    ws["B3"].fill = WARN_FILL
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 88
    ws.column_dimensions["C"].width = 56


def build() -> None:
    wb = Workbook()
    overview_sheet(wb)

    write_rows(
        wb.create_sheet("Admin"),
        ["row_id", "username", "mobile", "password", "email", "address", "userType", "notes"],
        [
            [
                1,
                "Admin",
                ADMIN_MOBILE,
                PASSWORD,
                "admin@cts.test",
                "admin@cts.test",
                "ADMIN",
                "Sign in with this account. Create other rows under this org.",
            ]
        ],
    )

    driver_rows = []
    cab_rows = []
    route_rows = []
    pop_rows = []
    batch_rows = []
    assignment_rows = []

    for d in range(DRIVER_COUNT):
        n = d + 1
        route = f"R{n:02d}"
        cab = f"GJ-06-D-{1000 + n}"
        batch = f"Batch-{n:02d}"
        driver_name = f"Driver {n}"
        mobile = str(DRIVER_BASE + d)
        morning = f"07:{30 + n:02d}:00"
        evening = f"18:{30 + n:02d}:00"
        driver_rows.append(
            [
                n,
                driver_name,
                driver_name,
                mobile,
                PASSWORD,
                f"driver{n}@cts.test",
                f"driver{n}@cts.test",
                "DRIVER",
                cab,
                batch,
                route,
                "QA login" if n == 1 else "",
            ]
        )
        cab_rows.append([n, cab, CAB_CAPACITY, route, 40 + n * 2, driver_name, batch])
        route_rows.append([n, route, f"Dummy corridor {n}"])
        batch_rows.append(
            [
                n,
                batch,
                morning,
                evening,
                "2026-08-01",
                "2026-12-31",
                driver_name,
                cab,
                route,
                ROSTER_PER_BATCH,
                COMING_PER_BATCH,
            ]
        )
        for p in range(POPS_PER_ROUTE):
            stop = STOP_NAMES[p]
            pop_name = f"{route}-{stop}"
            pop_rows.append(
                [
                    len(pop_rows) + 1,
                    pop_name,
                    route,
                    p + 1,
                    round(22.30 + n * 0.012 + p * 0.002, 6),
                    round(73.19 + n * 0.008 + p * 0.003, 6),
                ]
            )
        assignment_rows.append(
            [
                batch,
                driver_name,
                mobile,
                cab,
                route,
                ROSTER_PER_BATCH,
                " / ".join(f"{route}-{s}" for s in STOP_NAMES),
                COMING_PER_BATCH,
            ]
        )

    write_rows(
        wb.create_sheet("Drivers"),
        [
            "row_id",
            "username",
            "name",
            "mobile",
            "password",
            "email",
            "address",
            "userType",
            "cab_reg",
            "batch_name",
            "route_name",
            "notes",
        ],
        driver_rows,
    )
    write_rows(
        wb.create_sheet("Cabs"),
        ["row_id", "regNumber", "capacity", "route_name", "km", "driver_name", "batch_name"],
        cab_rows,
    )
    write_rows(
        wb.create_sheet("Routes"),
        ["row_id", "routeName", "notes"],
        route_rows,
    )
    write_rows(
        wb.create_sheet("POPs"),
        ["row_id", "pickUpPointName", "route_name", "inLine", "lat", "longitude"],
        pop_rows,
    )
    write_rows(
        wb.create_sheet("Batches"),
        [
            "row_id",
            "batchName",
            "batchTime",
            "returnTime",
            "startDate",
            "endDate",
            "driver_name",
            "cab_reg",
            "route_name",
            "planned_commuters",
            "planned_isComing",
        ],
        batch_rows,
    )
    write_rows(
        wb.create_sheet("Assignment_summary"),
        [
            "batch_name",
            "driver_name",
            "driver_mobile",
            "cab_reg",
            "route_name",
            "commuter_count",
            "pops",
            "isComing_true_count",
        ],
        assignment_rows,
    )

    commuter_rows = []
    stop_names = STOP_NAMES
    for i in range(COMMUTER_COUNT):
        if i < UG_COUNT:
            label = f"UG{i + 1}"
            college = "UG"
        else:
            pg_n = i - UG_COUNT + 1
            label = f"PG{pg_n}"
            college = "PG"
        batch_n = (i % DRIVER_COUNT) + 1
        pop_n = (i // DRIVER_COUNT) % POPS_PER_ROUTE
        coming = (i // DRIVER_COUNT) < COMING_PER_BATCH
        mobile = str(COMMUTER_BASE + i)
        email = f"{label.lower()}@cts.test"
        commuter_rows.append(
            [
                i + 1,
                label,
                mobile,
                PASSWORD,
                email,
                email,
                college,
                "COMMUTER",
                f"Batch-{batch_n:02d}",
                f"GJ-06-D-{1000 + batch_n}",
                f"R{batch_n:02d}",
                f"R{batch_n:02d}-{stop_names[pop_n]}",
                pop_n + 1,
                "TRUE" if coming else "FALSE",
                "QA login" if i == 0 else "",
            ]
        )

    write_rows(
        wb.create_sheet("Commuters"),
        [
            "row_id",
            "username",
            "mobile",
            "password",
            "email",
            "address",
            "collegeName",
            "userType",
            "batch_name",
            "cab_reg",
            "route_name",
            "pop_name",
            "pop_inLine",
            "isComing",
            "notes",
        ],
        commuter_rows,
    )

    counts = wb.create_sheet("Counts")
    write_rows(
        counts,
        ["sheet", "rows"],
        [
            ["Admin", 1],
            ["Drivers", DRIVER_COUNT],
            ["Cabs", DRIVER_COUNT],
            ["Routes", DRIVER_COUNT],
            ["POPs", DRIVER_COUNT * POPS_PER_ROUTE],
            ["Batches", DRIVER_COUNT],
            ["Commuters", COMMUTER_COUNT],
            ["Commuters UG", UG_COUNT],
            ["Commuters PG", PG_COUNT],
            ["Cab capacity (each)", CAB_CAPACITY],
            ["Roster per batch", ROSTER_PER_BATCH],
            ["isComing true", DRIVER_COUNT * COMING_PER_BATCH],
        ],
    )
    counts["B8"].fill = OK_FILL

    wb.save(OUT)
    print(f"Wrote {OUT} ({COMMUTER_COUNT} commuters)")


if __name__ == "__main__":
    build()
